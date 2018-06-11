from urllib.request import urlopen
import re
import os
from openpyxl import Workbook

url = "http://www.91zcp.com/info/notice-10/"
html = urlopen(url).read().decode("utf-8")
soup = BeautifulSoup(html, "lxml")
info = soup.find_all(class_="info_con")

banlist = info[0].find_all("p")[6]
inst = []
name = []
for text in banlist.strings:
    textlist = text.split()
    i = 0
    while i < len(textlist):
        if re.search(re.compile("\d{6}"), textlist[i]) and re.search(re.compile("\S+"), textlist[i+1]):
            inst.append(textlist[i])
            name.append(textlist[i+1])
        i += 2
wb = Workbook()
ws = wb.active
ws['A1'] = "INST"
ws['B1'] = "NAME"
for i in range(len(inst)):
    ws.cell(row=i+2, column=1, value=inst[i])
    ws.cell(row=i+2, column=2, value=name[i])
wb.save('banlist.xlsx')
